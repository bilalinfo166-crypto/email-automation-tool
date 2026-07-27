"""CRM API: domains, compliant scraping, contacts, suppression, campaigns
(with the review->approve compliance gate), queue, sending, and analytics."""
import re
import time
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from .database import get_db, SessionLocal
from .crm_models import Domain, Contact, Suppression, Campaign, QueueItem, CompanyProfile
from . import compliance, scraper, campaign_engine, replies, warmup

router = APIRouter(prefix="/crm", tags=["crm"])


# ---------------- company profile (one-time) ----------------

class ProfileIn(BaseModel):
    company_name: str
    website_url: str
    business_address: str
    sender_name: str
    reply_to_email: str


@router.get("/company-profile")
def get_company_profile(db: Session = Depends(get_db)):
    p = compliance.get_profile(db)
    if not p:
        return {"completed": False, "missing": [m for _, m in compliance.REQUIRED_PROFILE_FIELDS]}
    return {"completed": p.completed, "company_name": p.company_name, "website_url": p.website_url,
            "business_address": p.business_address, "sender_name": p.sender_name,
            "reply_to_email": p.reply_to_email, "missing": compliance.profile_missing(p)}


@router.post("/company-profile")
def save_company_profile(data: ProfileIn, db: Session = Depends(get_db)):
    if not compliance.valid_email(data.reply_to_email):
        raise HTTPException(400, "Reply-to email is not valid.")
    p = db.get(CompanyProfile, 1)
    if not p:
        p = CompanyProfile(id=1)
        db.add(p)
    for k, v in data.model_dump().items():
        setattr(p, k, v.strip())
    p.completed = len(compliance.profile_missing(p)) == 0
    db.commit()
    return {"completed": p.completed, "missing": compliance.profile_missing(p)}


# ---------------- domains + scraping ----------------

class DomainsIn(BaseModel):
    domains: list[str]
    mode: str = "vendor"


@router.post("/domains")
def upload_domains(data: DomainsIn, db: Session = Depends(get_db)):
    added = 0
    for raw in data.domains:
        d = raw.strip().lower().replace("www.", "")
        if not d:
            continue
        if not db.query(Domain).filter(Domain.domain == d).first():
            db.add(Domain(domain=d, mode=data.mode))
            added += 1
    db.commit()
    return {"added": added}


@router.post("/scrape")
def run_scrape(mode: str = "vendor", limit: int = 50, db: Session = Depends(get_db)):
    """Scrape this mode's pending domains for their OWN published business addresses.
    Dedups against existing contacts and skips suppressed emails."""
    pending = (db.query(Domain)
               .filter(Domain.status == "pending", Domain.mode == mode)
               .limit(limit).all())
    total_contacts = 0
    for dom in pending:
        result = scraper.extract_domain(dom.domain)
        found = 0
        for c in result["contacts"]:
            email = c["email"]
            if compliance.is_suppressed(db, email):
                continue
            if db.query(Contact).filter(Contact.email == email).first():
                continue  # dedup
            db.add(Contact(email=email, domain=c["domain"], mode=dom.mode,
                           source_url=c["source_url"],
                           role_based=c["role_based"], mx_ok=c["mx_ok"]))
            found += 1
        dom.status = result["status"] if not result["contacts"] else "scraped"
        dom.contacts_found = found
        total_contacts += found
        db.commit()
    return {"domains_processed": len(pending), "new_contacts": total_contacts}


@router.get("/contacts")
def list_contacts(mode: str = "", db: Session = Depends(get_db)):
    q = db.query(Contact)
    if mode:
        q = q.filter(Contact.mode == mode)
    rows = q.all()
    return [{"id": c.id, "email": c.email, "domain": c.domain, "mode": c.mode,
             "role_based": c.role_based, "mx_ok": c.mx_ok, "status": c.status,
             "source_url": c.source_url} for c in rows]


# ---------------- suppression ----------------

class SuppressIn(BaseModel):
    email: str
    reason: str = "manual"


@router.post("/suppression")
def add_suppress(data: SuppressIn, db: Session = Depends(get_db)):
    if not compliance.valid_email(data.email):
        raise HTTPException(400, "Invalid email")
    compliance.add_suppression(db, data.email, data.reason)
    return {"suppressed": data.email.lower()}


@router.get("/suppression")
def list_suppress(db: Session = Depends(get_db)):
    return [{"email": s.email, "reason": s.reason} for s in db.query(Suppression).all()]


# ---------------- campaigns + compliance gate ----------------

class CampaignIn(BaseModel):
    name: str
    mode: str = "vendor"
    subject: str = ""
    from_name: str = ""
    company: str = ""
    postal_address: str = ""
    reason_for_contact: str = ""
    body_html: str = ""
    unsubscribe_url: str = ""
    per_sender_daily_cap: int = 100
    min_delay_sec: int = 25
    max_delay_sec: int = 60


@router.post("/campaigns")
def create_campaign(data: CampaignIn, db: Session = Depends(get_db)):
    c = Campaign(**data.model_dump())
    db.add(c); db.commit(); db.refresh(c)
    return {"id": c.id, "status": c.status, "mode": c.mode}


@router.get("/campaigns")
def list_campaigns(mode: str = "", db: Session = Depends(get_db)):
    q = db.query(Campaign)
    if mode:
        q = q.filter(Campaign.mode == mode)
    return [{"id": c.id, "name": c.name, "mode": c.mode, "status": c.status,
             "lawful_basis_confirmed": c.lawful_basis_confirmed} for c in q.all()]


@router.post("/campaigns/{cid}/review")
def review_campaign(cid: int, db: Session = Depends(get_db)):
    """Campaign review step: reports what's missing before it can be approved."""
    c = db.get(Campaign, cid)
    if not c:
        raise HTTPException(404, "Campaign not found")
    prof_miss = compliance.profile_missing(compliance.get_profile(db))
    miss = compliance.missing_fields(c)
    ready = not miss and not prof_miss
    if ready:
        c.status = "pending_review"; db.commit()
    return {"ready_to_approve": ready, "missing_campaign_fields": miss,
            "missing_company_profile": prof_miss,
            "needs_lawful_basis_confirmation": not c.lawful_basis_confirmed}


class ApproveIn(BaseModel):
    lawful_basis_confirmed: bool = False


@router.post("/campaigns/{cid}/approve")
def approve_campaign(cid: int, data: ApproveIn, db: Session = Depends(get_db)):
    """Hard gate: approval requires all fields present AND explicit lawful-basis confirmation."""
    c = db.get(Campaign, cid)
    if not c:
        raise HTTPException(404, "Campaign not found")
    miss = compliance.missing_fields(c)
    if miss:
        raise HTTPException(400, "Cannot approve — missing: " + ", ".join(miss))
    if not data.lawful_basis_confirmed:
        raise HTTPException(400, "You must confirm you have a lawful basis / permission to contact these businesses.")
    c.lawful_basis_confirmed = True
    c.status = "approved"
    db.commit()
    return {"id": c.id, "status": c.status}


@router.post("/campaigns/{cid}/build-queue")
def build_queue(cid: int, db: Session = Depends(get_db)):
    try:
        return campaign_engine.build_queue(db, cid)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/campaigns/{cid}/send-batch")
def send_batch(cid: int, batch_size: int = 20, db: Session = Depends(get_db)):
    try:
        return campaign_engine.send_batch(db, cid, batch_size)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/campaigns/{cid}/analytics")
def campaign_analytics(cid: int, db: Session = Depends(get_db)):
    return campaign_engine.analytics(db, cid)


# ---------------- replies + opt-out ----------------

@router.post("/replies/scan")
def scan_replies(db: Session = Depends(get_db)):
    """Read OAuth senders' inboxes; log replies and auto-handle opt-outs."""
    return replies.scan_replies(db)


class ManualReplyIn(BaseModel):
    email: str
    text: str


@router.post("/replies/manual")
def manual_reply(data: ManualReplyIn, db: Session = Depends(get_db)):
    """Record a reply by hand (e.g. for app-password senders). Applies opt-out wording."""
    contact = db.query(Contact).filter(Contact.email == data.email.lower()).first()
    if not contact:
        raise HTTPException(404, "Contact not found")
    cid = replies._campaign_for(db, contact.id)
    status = replies.apply_reply(db, contact, cid, 0, data.text)
    return {"email": contact.email, "status": status}


@router.get("/campaigns/{cid}/recipients")
def campaign_recipients(cid: int, db: Session = Depends(get_db)):
    """Live status per recipient for the campaign table."""
    label = {"new": "Queued", "queued": "Queued", "sent": "Sent", "failed": "Failed",
             "replied": "Replied", "unsubscribed": "Unsubscribed", "not_interested": "Not Interested"}
    out = []
    items = db.query(QueueItem).filter(QueueItem.campaign_id == cid).all()
    for it in items:
        ct = db.get(Contact, it.contact_id)
        if not ct:
            continue
        # contact-level status (replied/unsub/not_interested) wins over queue status
        status = ct.status if ct.status in ("replied", "unsubscribed", "not_interested") else it.status
        out.append({"email": ct.email, "status": label.get(status, status.title()),
                    "sent_at": it.sent_at.isoformat() if it.sent_at else None})
    return out


# ---------------- warmup ----------------

@router.get("/warmup/status")
def warmup_status(db: Session = Depends(get_db)):
    return warmup.warmup_status(db)


@router.post("/warmup/run")
def warmup_run(max_sends: int = 10, db: Session = Depends(get_db)):
    """Send a small batch of warmup emails between the user's own connected inboxes."""
    return warmup.run_warmup(db, max_sends)


# ---------------- compliance dashboard + full stats ----------------

@router.get("/dashboard")
def compliance_dashboard(mode: str = "", db: Session = Depends(get_db)):
    from .crm_models import EventLog, ScraperJob, ScraperResult
    from .database import Sender
    try:
        return _compliance_dashboard_impl(mode, db)
    except Exception as e:
        import traceback
        print(f"[dashboard] ERROR for mode={mode}: {e}")
        traceback.print_exc()
        # Minimal safe payload so the UI still renders
        return {"mode": mode or "all", "company_profile_completed": False,
                "statuses": {"Sent":0,"Opened":0,"Replied":0,"Failed":0,
                             "Unsubscribed":0,"Not Interested":0},
                "senders": [], "error": str(e)}


def _compliance_dashboard_impl(mode, db):
    from .crm_models import EventLog, ScraperJob, ScraperResult
    from .database import Sender
    cq = db.query(Campaign)
    if mode:
        cq = cq.filter(Campaign.mode == mode)
    campaign_ids = [c.id for c in cq.all()]

    totals = {"sent": 0, "opened": 0, "replied": 0, "failed": 0, "unsubscribed": 0, "not_interested": 0}
    eq = db.query(EventLog)
    if mode:
        eq = eq.filter(EventLog.campaign_id.in_(campaign_ids or [-1]))
    for ev in eq.all():
        if ev.type in totals:
            totals[ev.type] += 1

    contacts_q = db.query(Contact)
    if mode:
        contacts_q = contacts_q.filter(Contact.mode == mode)

    # scraper stats (mode-aware)
    scraper_q = db.query(ScraperJob)
    if mode:
        scraper_q = scraper_q.filter(ScraperJob.mode == mode)
    scraper_jobs_all = scraper_q.all()
    total_scraped_emails = sum(j.emails_found for j in scraper_jobs_all)
    total_domains_scraped = sum(j.total for j in scraper_jobs_all)

    # per-sender stats
    senders = db.query(Sender).all()
    sender_stats = []
    for s in senders:
        sender_sent = db.query(EventLog).filter(EventLog.sender_id == s.id, EventLog.type == "sent").count()
        sender_failed = db.query(EventLog).filter(EventLog.sender_id == s.id, EventLog.type == "failed").count()
        sender_replied = db.query(EventLog).filter(EventLog.sender_id == s.id, EventLog.type == "replied").count()
        sender_stats.append({
            "id": s.id, "email": s.email, "name": s.name or s.email.split("@")[0],
            "method": s.method, "health": s.health, "status": s.status,
            "sent": sender_sent, "failed": sender_failed, "replied": sender_replied,
            "sent_today": s.sent_today, "total_sent": s.total_sent, "daily_cap": s.daily_cap,
        })

    prof = compliance.get_profile(db)
    return {
        "mode": mode or "all",
        "company_profile_completed": bool(prof and prof.completed),
        "statuses": {
            "Sent": totals["sent"], "Opened": totals["opened"], "Replied": totals["replied"],
            "Failed": totals["failed"], "Unsubscribed": totals["unsubscribed"],
            "Not Interested": totals["not_interested"],
        },
        "contacts": contacts_q.count(),
        "suppressed": db.query(Suppression).count(),
        "campaigns": len(campaign_ids),
        "approved_campaigns": cq.filter(Campaign.status.in_(["approved", "sending", "completed"])).count(),
        "unsubscribe_rate": round(totals["unsubscribed"] / totals["sent"] * 100, 2) if totals["sent"] else 0.0,
        "scraper": {
            "total_jobs": len(scraper_jobs_all),
            "total_domains": total_domains_scraped,
            "total_emails_found": total_scraped_emails,
        },
        "sender_stats": sender_stats,
    }


@router.get("/stats/scraper-history")
def scraper_history(mode: str = "", db: Session = Depends(get_db)):
    """All scraper jobs with their results — for the expandable stats view."""
    from .crm_models import ScraperJob, ScraperResult
    q = db.query(ScraperJob)
    if mode:
        q = q.filter(ScraperJob.mode == mode)
    jobs = q.order_by(ScraperJob.id.desc()).all()
    out = []
    for j in jobs:
        results = db.query(ScraperResult).filter(ScraperResult.job_id == j.id).all()
        out.append({
            "id": j.id, "name": j.name, "mode": j.mode, "status": j.status,
            "total": j.total, "done": j.done, "emails_found": j.emails_found,
            "created_at": j.created_at.isoformat(),
            "emails": [{"email": r.email, "domain": r.domain, "email_type": r.email_type,
                        "confidence": r.confidence, "source_url": r.source_url}
                       for r in results],
        })
    return out


@router.get("/stats/sender-history")
def sender_history(db: Session = Depends(get_db)):
    """Per-sender detailed send log — for expandable sender stats."""
    from .crm_models import EventLog
    from .database import Sender
    senders = db.query(Sender).all()
    out = []
    for s in senders:
        events = (db.query(EventLog).filter(EventLog.sender_id == s.id)
                  .order_by(EventLog.id.desc()).limit(200).all())
        out.append({
            "id": s.id, "email": s.email, "name": s.name or s.email.split("@")[0],
            "method": s.method, "health": s.health, "status": s.status,
            "total_sent": s.total_sent, "sent_today": s.sent_today, "daily_cap": s.daily_cap,
            "events": [{"type": e.type, "campaign_id": e.campaign_id,
                        "created_at": e.created_at.isoformat()} for e in events],
        })
    return out


# ============ OUTREACH SHEET (Campaign Send List) ============

@router.post("/outreach/add-from-scraper")
def add_scraper_results_to_outreach(mode: str = "vendor", job_id: int = 0, db: Session = Depends(get_db)):
    """Add scraped emails to this mode's send list.

    Safe to click ANY TIME — including while the scraper is still running.
    Whatever has been found so far gets added; clicking again later adds only
    the new ones. Never creates duplicates and never re-adds unsubscribed
    addresses. Data is per-mode, so vendor/client/blog lists stay separate.
    """
    from .crm_models import ScraperResult, ScraperJob, OutreachEntry

    if job_id > 0:
        # Only allow a job that belongs to THIS mode (keeps modes separate)
        job = db.get(ScraperJob, job_id)
        if not job or (job.mode or "vendor") != mode:
            return {"added": 0, "total": db.query(OutreachEntry).filter(
                OutreachEntry.mode == mode).count(),
                "error": "Job not found for this dashboard."}
        results = db.query(ScraperResult).filter(ScraperResult.job_id == job_id).all()
    else:
        job_ids = [j.id for j in db.query(ScraperJob).filter(ScraperJob.mode == mode).all()]
        results = db.query(ScraperResult).filter(ScraperResult.job_id.in_(job_ids or [-1])).all()

    # Load ALL existing emails for this mode in ONE query (was one query per
    # email — far too slow for thousands of results).
    existing = {e[0].strip().lower() for e in db.query(OutreachEntry.email).filter(
        OutreachEntry.mode == mode).all() if e[0]}
    # Same for the unsubscribe list — one query, not one per email. Querying per
    # email made this endpoint hang for minutes on big lists (and it competes
    # with the scraper for the database).
    from .crm_models import Suppression
    suppressed = {s[0].strip().lower() for s in db.query(Suppression.email).all() if s[0]}

    added = skipped_dup = skipped_suppressed = 0
    new_rows = []
    for r in results:
        email = (r.email or "").strip().lower()
        if not email or "@" not in email:
            continue
        if email in existing:          # already in list (or seen earlier in this batch)
            skipped_dup += 1
            continue
        if email in suppressed:        # unsubscribed — never re-add
            skipped_suppressed += 1
            continue
        new_rows.append(OutreachEntry(mode=mode, email=email, domain=r.domain,
            email_type=r.email_type, confidence=r.confidence, source_url=r.source_url))
        existing.add(email)            # prevents duplicates within this batch too
        added += 1

    # Insert in chunks so one huge transaction doesn't hold the write lock
    # while the scraper is running.
    CHUNK = 500
    for i in range(0, len(new_rows), CHUNK):
        chunk = new_rows[i:i + CHUNK]
        try:
            db.add_all(chunk)
            db.commit()
        except Exception:
            # A duplicate slipped past the in-memory check (e.g. two imports at
            # once). The unique index rejected the whole chunk, so retry row by
            # row and silently skip the ones that already exist.
            db.rollback()
            for row in chunk:
                try:
                    db.add(row); db.commit()
                except Exception:
                    db.rollback()
                    added -= 1
                    skipped_dup += 1
    if not new_rows:
        db.commit()

    total = db.query(OutreachEntry).filter(OutreachEntry.mode == mode).count()
    return {"added": added, "total": total,
            "skipped_duplicates": skipped_dup,
            "skipped_unsubscribed": skipped_suppressed}


@router.get("/outreach/list")
def list_outreach(mode: str = "vendor", page: int = 1, limit: int = 100, status: str = "", db: Session = Depends(get_db)):
    """Paginated outreach sheet with live status and optional filter."""
    from .crm_models import OutreachEntry
    q = db.query(OutreachEntry).filter(OutreachEntry.mode == mode)
    if status:
        q = q.filter(OutreachEntry.status == status)
    total = q.count()
    entries = q.order_by(OutreachEntry.id.desc()).offset((page-1)*limit).limit(limit).all()
    return {
        "total": total, "page": page, "limit": limit,
        "entries": [{"id":e.id,"email":e.email,"domain":e.domain,"email_type":e.email_type,
                     "confidence":e.confidence,"source_url":e.source_url,"status":e.status,
                     "sent_at":e.sent_at.isoformat() if e.sent_at else None,
                     "opened_at":e.opened_at.isoformat() if e.opened_at else None,
                     "replied_at":e.replied_at.isoformat() if e.replied_at else None,
                     "sender_email":e.sender_email,"subject":e.subject,
                     "created_at":e.created_at.isoformat()} for e in entries]
    }


@router.get("/outreach/stats")
def outreach_stats(mode: str = "vendor", db: Session = Depends(get_db)):
    """Live stats for the outreach sheet. Hardened so a single bad row can
    never 500 the whole endpoint."""
    from .crm_models import OutreachEntry
    try:
        q = db.query(OutreachEntry).filter(OutreachEntry.mode == mode)
        total = q.count()
        sent = q.filter(OutreachEntry.status.in_(["sent","opened","replied"])).count()
        opened = q.filter(OutreachEntry.status.in_(["opened","replied"])).count()
        replied = q.filter(OutreachEntry.status == "replied").count()
        bounced = q.filter(OutreachEntry.status == "bounced").count()
        pending = q.filter(OutreachEntry.status == "pending").count()
        return {"total":total,"pending":pending,"sent":sent,"opened":opened,
                "replied":replied,"bounced":bounced,
                "open_rate":round(opened/max(1,sent)*100,1),
                "reply_rate":round(replied/max(1,sent)*100,1)}
    except Exception as e:
        import traceback
        print(f"[outreach_stats] ERROR for mode={mode}: {e}")
        traceback.print_exc()
        # Return zeros instead of 500 so the dashboard still renders
        return {"total":0,"pending":0,"sent":0,"opened":0,"replied":0,
                "bounced":0,"open_rate":0.0,"reply_rate":0.0,"error":str(e)}


@router.get("/outreach/export")
def export_outreach(mode: str = "vendor", format: str = "csv", db: Session = Depends(get_db)):
    """Export outreach sheet as CSV or Excel."""
    from .crm_models import OutreachEntry
    from fastapi.responses import Response
    entries = db.query(OutreachEntry).filter(OutreachEntry.mode == mode).order_by(OutreachEntry.id).all()
    if format == "xlsx":
        from openpyxl import Workbook
        from io import BytesIO
        wb = Workbook(); ws = wb.active
        ws.append(["Email","Domain","Type","Confidence","Status","Sent","Opened","Replied","Source","Added"])
        for e in entries:
            ws.append([e.email,e.domain,e.email_type,e.confidence,e.status,
                       str(e.sent_at or ""),str(e.opened_at or ""),str(e.replied_at or ""),
                       e.source_url,str(e.created_at)])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return Response(content=buf.read(),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition":"attachment; filename=outreach.xlsx"})
    # CSV
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Email","Domain","Type","Confidence","Status","Sent","Opened","Replied","Source","Added"])
    for e in entries:
        w.writerow([e.email,e.domain,e.email_type,e.confidence,e.status,
                    e.sent_at or "",e.opened_at or "",e.replied_at or "",e.source_url,e.created_at])
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition":"attachment; filename=outreach.csv"})


# ============ AUTO-CAMPAIGN BUILDER ============

@router.post("/outreach/build-campaigns")
def build_campaigns(mode: str = "vendor", db: Session = Depends(get_db)):
    """Auto-create campaigns from pending outreach emails.
    Each sender gets one campaign with emails up to their daily_cap."""
    from .crm_models import OutreachEntry
    from .database import Sender

    senders = db.query(Sender).filter(Sender.mode == mode).all()
    if not senders:
        return {"error": "No senders in this mode. Add senders first.", "campaigns": []}

    pending = db.query(OutreachEntry).filter(
        OutreachEntry.mode == mode, OutreachEntry.status == "pending"
    ).order_by(OutreachEntry.id).all()

    if not pending:
        return {"error": "No pending emails in send list.", "campaigns": []}

    created = []
    idx = 0
    for sender in senders:
        if idx >= len(pending):
            break
        cap = sender.daily_cap or 150
        batch = pending[idx:idx + cap]
        idx += cap

        # Create campaign
        camp = Campaign(
            mode=mode,
            name=f"Campaign — {sender.email} ({len(batch)} emails)",
            status="ready",
        )
        db.add(camp)
        db.commit()
        db.refresh(camp)

        # Link emails to this campaign
        for entry in batch:
            entry.status = "queued"
            entry.sender_email = sender.email
            db.add(QueueItem(
                campaign_id=camp.id,
                contact_id=0,
                sender_id=sender.id,
                email=entry.email,
                subject="",
                body_html="",
            ))
        db.commit()

        created.append({
            "campaign_id": camp.id,
            "sender": sender.email,
            "emails": len(batch),
            "status": "ready"
        })

    remaining = len(pending) - idx
    return {"campaigns": created, "remaining_pending": remaining,
            "total_queued": idx}


@router.get("/outreach/campaigns")
def list_outreach_campaigns(mode: str = "vendor", db: Session = Depends(get_db)):
    """List all campaigns with stats."""
    from .crm_models import OutreachEntry
    camps = db.query(Campaign).filter(Campaign.mode == mode).order_by(Campaign.id.desc()).all()
    out = []
    for c in camps:
        # Count statuses from outreach entries linked to this campaign's sender
        total_queued = db.query(QueueItem).filter(QueueItem.campaign_id == c.id).count()
        sent = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "sent").count()
        opened = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "opened").count()
        replied = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "replied").count()
        failed = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "failed").count()
        out.append({
            "id": c.id, "name": c.name, "status": c.status, "mode": c.mode,
            "total": total_queued, "sent": sent, "opened": opened,
            "replied": replied, "failed": failed,
            "created_at": c.created_at.isoformat() if c.created_at else ""
        })
    return out


# ============ SEND ENGINE ============

@router.post("/send/start")
def start_sending(campaign_id: int, mode: str = "vendor", 
                  emails_per_batch: int = 10, delay_seconds: int = 60,
                  scheduled_time: str = None, db: Session = Depends(get_db)):
    """Start sending emails for a campaign with rate control and optional scheduling."""
    from . import send_engine
    result = send_engine.start_campaign_send(
        campaign_id=campaign_id, mode=mode,
        emails_per_batch=emails_per_batch,
        delay_seconds=delay_seconds,
        scheduled_time=scheduled_time
    )
    return result


@router.post("/send/stop")
def stop_sending(campaign_id: int, db: Session = Depends(get_db)):
    """Stop a running campaign."""
    from . import send_engine
    return send_engine.stop_campaign(campaign_id)


# ============ TEMPLATES ============

@router.get("/templates")
def list_templates(mode: str = "client"):
    """Return outreach templates for the mode:
    vendor = asking sites for a paid guest post
    blog   = pitching prospects found via blog research (names their article)
    client = pitching our guest-posting service
    """
    if mode == "vendor":
        from .vendor_templates import VENDOR_TEMPLATES
        return [{"id": i+1, "subject": t["subject"], "body": t["body"]} for i, t in enumerate(VENDOR_TEMPLATES)]
    if mode == "blog":
        from .blog_templates import BLOG_TEMPLATES
        return [{"id": i+1, "subject": t["subject"], "body": t["body"]} for i, t in enumerate(BLOG_TEMPLATES)]
    from .email_templates import TEMPLATES
    return [{"id": i+1, "subject": t["subject"], "body": t["body"]} for i, t in enumerate(TEMPLATES)]


@router.post("/outreach/import-csv")
def import_csv_to_outreach(mode: str = "client", db: Session = Depends(get_db)):
    """Import emails from scraper_6__3_.csv in the working directory."""
    import csv, os
    from .crm_models import OutreachEntry
    # Try multiple possible locations
    paths = ["scraper_6__3_.csv", "app/scraper_6__3_.csv", "../scraper_6__3_.csv",
             os.path.expanduser("~/Desktop/files/scraper_6__3_.csv"),
             "C:/Users/Dell/Desktop/files/scraper_6__3_.csv",
             "C:/Users/Dell/Downloads/scraper_6__3_.csv"]
    csv_path = None
    for p in paths:
        if os.path.exists(p):
            csv_path = p; break
    if not csv_path:
        return {"error": "CSV file not found. Place scraper_6__3_.csv in C:\\Users\\Dell\\Desktop\\files\\", "paths_checked": paths}
    added = 0
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            email = (row.get('Email') or '').strip()
            domain = (row.get('Domain') or '').strip()
            if not email or '@' not in email: continue
            exists = db.query(OutreachEntry).filter(OutreachEntry.email == email, OutreachEntry.mode == mode).first()
            if not exists:
                db.add(OutreachEntry(mode=mode, email=email, domain=domain,
                    email_type=row.get('Email Type','domain_email'),
                    confidence=row.get('Confidence','medium'),
                    source_url=row.get('Source URL',''), status='pending'))
                added += 1
    db.commit()
    total = db.query(OutreachEntry).filter(OutreachEntry.mode == mode).count()
    return {"added": added, "total": total, "file": csv_path}


@router.post("/outreach/verify")
def verify_outreach_emails(mode: str = "vendor", db: Session = Depends(get_db)):
    """Verify all pending emails (MX check). Marks invalid ones as bounced."""
    from .crm_models import OutreachEntry
    from .email_verify import quick_verify
    pending = db.query(OutreachEntry).filter(
        OutreachEntry.mode == mode, OutreachEntry.status == "pending").all()
    valid = 0; invalid = 0
    for entry in pending:
        is_valid, reason = quick_verify(entry.email)
        if not is_valid:
            entry.status = "bounced"
            invalid += 1
        else:
            valid += 1
    db.commit()
    return {"checked": len(pending), "valid": valid, "invalid": invalid}


@router.get("/senders/activity")
def sender_activity(mode: str = "vendor", page: int = 1, limit: int = 100,
                    sender: str = "", db: Session = Depends(get_db)):
    """Live sending history — which email sent via which sender, status, dates."""
    from .crm_models import OutreachEntry
    q = db.query(OutreachEntry).filter(
        OutreachEntry.mode == mode,
        OutreachEntry.status.in_(["sent", "opened", "replied", "bounced"])
    )
    if sender:
        q = q.filter(OutreachEntry.sender_email == sender)
    total = q.count()
    entries = q.order_by(OutreachEntry.sent_at.desc()).offset((page-1)*limit).limit(limit).all()
    return {
        "total": total, "page": page,
        "entries": [{"email": e.email, "domain": e.domain,
                     "sender_email": e.sender_email or "—",
                     "status": e.status, "subject": e.subject,
                     "sent_at": e.sent_at.isoformat() if e.sent_at else None,
                     "opened_at": e.opened_at.isoformat() if e.opened_at else None,
                     "replied_at": e.replied_at.isoformat() if e.replied_at else None
                     } for e in entries]
    }


# ============ WARMUP ENGINE ============

@router.post("/warmup/start")
def start_warmup_engine(mode: str = "client", interval_minutes: int = 90):
    """Start real-time warmup — senders email each other, rescue from spam, reply."""
    from . import warmup_engine
    return warmup_engine.start_warmup(mode, interval_minutes)


@router.post("/warmup/stop")
def stop_warmup_engine():
    from . import warmup_engine
    return warmup_engine.stop_warmup()


@router.post("/warmup/run-once")
def run_warmup_once(mode: str = "client"):
    """Run one warmup cycle immediately (for testing)."""
    from . import warmup_engine
    return warmup_engine.run_warmup_cycle(mode)


@router.get("/warmup/engine-status")
def get_warmup_engine_status(mode: str = "client", db: Session = Depends(get_db)):
    from . import warmup_engine
    from .database import Sender
    status = warmup_engine.warmup_status()
    pool = warmup_engine.pool_info(db)
    senders = db.query(Sender).filter(Sender.mode == mode, Sender.warmup == True).all()
    return {
        "running": status["running"],
        "active_senders": len(senders),
        "pool_size": pool["pool_size"],
        "pool_domains": pool["domains"],
        "can_warmup": pool["can_warmup"],
        "domain_breakdown": pool["domain_breakdown"],
        "senders": [{"email": s.email, "warmup_sent_today": s.warmup_sent_today or 0,
                     "total_sent": s.total_sent or 0} for s in senders]
    }


# ============ FULL CAMPAIGN SYSTEM (autopilot, scheduling, sender selection) ============

@router.post("/campaign/create")
def create_campaign(
    name: str, mode: str = "vendor",
    sender_emails: str = "",       # comma-separated: "a@x.com,b@y.com"
    emails_per_batch: int = 10,
    delay_seconds: int = 30,
    min_delay: int = 0, max_delay: int = 0,   # random gap range
    total_target: int = 0,          # 0 = all pending
    scheduled_time: str = "",       # ISO datetime, empty = manual
    autopilot: bool = False,
    db: Session = Depends(get_db)
):
    """Create a campaign with full control: which senders, timing, autopilot."""
    from .crm_models import OutreachEntry
    pending = db.query(OutreachEntry).filter(
        OutreachEntry.mode == mode, OutreachEntry.status == "pending").count()
    if pending == 0:
        return {"error": "No pending emails in send list."}
    target = total_target if total_target > 0 else pending
    # Random gap range (default: build from delay_seconds)
    if min_delay <= 0:
        min_delay = max(10, int(delay_seconds * 0.5))
    if max_delay <= 0:
        max_delay = int(delay_seconds * 1.35)

    camp = Campaign(
        name=name, mode=mode, status="scheduled" if scheduled_time else "ready",
        sender_emails=sender_emails, emails_per_batch=emails_per_batch,
        delay_seconds=delay_seconds, min_delay_sec=min_delay, max_delay_sec=max_delay,
        scheduled_time=scheduled_time,
        autopilot=autopilot, total_target=target,
    )
    db.add(camp); db.commit(); db.refresh(camp)
    return {"campaign_id": camp.id, "name": camp.name, "target": target,
            "status": camp.status, "autopilot": autopilot,
            "gap": f"{min_delay}-{max_delay}s random"}


@router.get("/campaign/list")
def list_all_campaigns(mode: str = "vendor", db: Session = Depends(get_db)):
    """List all campaigns with full details."""
    camps = db.query(Campaign).filter(Campaign.mode == mode).order_by(Campaign.id.desc()).all()
    from .crm_models import EventLog, QueueItem
    out = []
    for c in camps:
        sent = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "sent").count()
        opened = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "opened").count()
        replied = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "replied").count()
        failed = db.query(EventLog).filter(EventLog.campaign_id == c.id, EventLog.type == "failed").count()
        out.append({
            "id": c.id, "name": c.name, "status": c.status,
            "senders": c.sender_emails or "all", "target": c.total_target,
            "sent": sent, "opened": opened, "replied": replied, "failed": failed,
            "batch": c.emails_per_batch, "delay": c.delay_seconds,
            "scheduled_time": c.scheduled_time, "autopilot": c.autopilot,
            "created_at": c.created_at.isoformat() if c.created_at else ""
        })
    return out


@router.post("/campaign/{campaign_id}/start")
def start_campaign(campaign_id: int, db: Session = Depends(get_db)):
    """Start (or schedule) a specific campaign."""
    from . import send_engine
    camp = db.get(Campaign, campaign_id)
    if not camp:
        return {"error": "Campaign not found"}
    result = send_engine.start_campaign_send(
        campaign_id=campaign_id, mode=camp.mode,
        emails_per_batch=camp.emails_per_batch,
        delay_seconds=camp.delay_seconds,
        min_delay=camp.min_delay_sec or 0,
        max_delay=camp.max_delay_sec or 0,
        scheduled_time=camp.scheduled_time or None,
        sender_filter=camp.sender_emails or None,
        total_target=camp.total_target,
        autopilot=camp.autopilot,
    )
    return result


@router.post("/campaign/{campaign_id}/stop")
def stop_campaign_ep(campaign_id: int, db: Session = Depends(get_db)):
    from . import send_engine
    return send_engine.stop_campaign(campaign_id)


@router.delete("/campaign/{campaign_id}")
def delete_campaign(campaign_id: int, db: Session = Depends(get_db)):
    camp = db.get(Campaign, campaign_id)
    if camp:
        db.delete(camp); db.commit()
    return {"deleted": campaign_id}


# ============ BLOG RESEARCH (client-hunting via external links) ============

_blog_threads = {}
_blog_stop = {}


@router.post("/blog/backfill-articles")
def blog_backfill_articles(db: Session = Depends(get_db)):
    """Fill in the source site + article for blog prospects saved earlier.

    Blog research always recorded which site and article a link came from — it
    just wasn't copied onto the outreach row until recently. This copies it
    across, so older prospects get the same personalised email as new ones.
    """
    from .crm_models import OutreachEntry, BlogResearchLink

    todo = db.query(OutreachEntry).filter(
        OutreachEntry.mode == "blog",
        (OutreachEntry.ref_article == "") | (OutreachEntry.ref_article.is_(None)),
    ).all()
    if not todo:
        return {"updated": 0, "still_missing": 0,
                "note": "Every blog prospect already has its article."}

    # Load the link records once and index them (fast, no query per row)
    by_email, by_domain = {}, {}
    for lk in db.query(BlogResearchLink).filter(
            BlogResearchLink.source_article != "").all():
        if lk.email:
            by_email.setdefault(lk.email.strip().lower(), lk)
        if lk.target_domain:
            by_domain.setdefault(lk.target_domain.replace("www.", "").lower(), lk)

    updated = 0
    for e in todo:
        lk = by_email.get((e.email or "").strip().lower())
        if lk is None:
            lk = by_domain.get((e.domain or "").replace("www.", "").lower())
        if lk is None:
            continue
        e.ref_site = lk.source_site or ""
        e.ref_article = lk.source_article or ""
        updated += 1
    db.commit()

    missing = db.query(OutreachEntry).filter(
        OutreachEntry.mode == "blog",
        (OutreachEntry.ref_article == "") | (OutreachEntry.ref_article.is_(None)),
    ).count()
    print(f"[BlogResearch] Backfilled article details for {updated} prospect(s).")
    return {"updated": updated, "still_missing": missing,
            "note": "Older blog prospects now carry the site and article we found "
                    "them in, so their emails are personalised too."}


@router.post("/senders/recalc")
def recalc_sender_stats(db: Session = Depends(get_db)):
    """Rebuild each sender's counters from the real sending history.

    Deleting and re-adding a sender wipes its own counters (sent/replies), even
    though every email it sent is still on record. This recomputes them from
    those records, so the numbers come back.
    """
    from .crm_models import OutreachEntry
    from .database import Sender
    from datetime import datetime, timedelta

    midnight = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    out = []
    for s in db.query(Sender).all():
        base = db.query(OutreachEntry).filter(OutreachEntry.sender_email == s.email)
        total = base.filter(OutreachEntry.status.in_(
            ["sent", "opened", "replied", "bounced"])).count()
        replied = base.filter(OutreachEntry.status == "replied").count()
        bounced = base.filter(OutreachEntry.status == "bounced").count()
        today = base.filter(OutreachEntry.sent_at >= midnight).count()

        before = s.total_sent or 0
        s.total_sent = total
        s.sent_today = today
        s.replies = replied
        s.failed = bounced
        out.append({"sender": s.email, "was": before, "now": total,
                    "sent_today": today, "replies": replied, "bounced": bounced})
    db.commit()
    print(f"[Senders] Recalculated stats for {len(out)} sender(s).")
    return {"senders": out,
            "note": "Counters rebuilt from the emails actually on record."}


@router.post("/labels/apply")
def labels_apply(limit: int = 300, db: Session = Depends(get_db)):
    """Apply Gmail labels to sent messages that don't have them yet."""
    from . import gmail_labels
    try:
        return gmail_labels.label_pending(db, limit=limit)
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"labelled": 0, "error": str(e)}


@router.get("/settings/own-domains")
def get_own_domains(db: Session = Depends(get_db)):
    """Domains treated as ours, so blog research never offers them as prospects."""
    from . import blog_research
    from .database import Sender
    auto = set()
    for snd in db.query(Sender).all():
        if snd.email and "@" in snd.email:
            auto.add(snd.email.split("@")[-1].lower())
    return {"domains": sorted(blog_research._OWN_DOMAINS),
            "brands": sorted(blog_research._OWN_BRANDS),
            "from_senders": sorted(auto),
            "note": "Taken from your senders automatically. Add anything else "
                    "you own with POST /crm/settings/own-domains?domains=a.com,b.com"}


@router.post("/settings/own-domains")
def set_own_domains_ep(domains: str = "", brands: str = ""):
    """Add your own sites so they're never picked up as prospects."""
    from . import blog_research
    extra = [d.strip() for d in re.split(r"[,\n ]+", domains) if d.strip()]
    extra_brands = [b.strip() for b in re.split(r"[,\n ]+", brands) if b.strip()]
    current = set(blog_research._OWN_DOMAINS) | set(extra)
    current_brands = set(blog_research._OWN_BRANDS) | set(extra_brands)
    blog_research.set_own_domains(current, brands=current_brands)
    return {"domains": sorted(blog_research._OWN_DOMAINS),
            "brands": sorted(blog_research._OWN_BRANDS)}


@router.post("/deals/backfill")
def deals_backfill(mode: str = "", db: Session = Depends(get_db)):
    """Create deal rows for everyone who has already replied.

    Rows normally appear as replies are detected. This fills in the ones that
    were marked as replied before the deals sheet existed, so nothing is
    missing. Prices are filled in on the next reply check, when their
    conversation is read.
    """
    from .crm_models import OutreachEntry, Deal
    from . import reply_tracker

    q = db.query(OutreachEntry).filter(OutreachEntry.status == "replied")
    if mode:
        q = q.filter(OutreachEntry.mode == mode)
    rows = q.all()

    created = 0
    for e in rows:
        email = (e.email or "").strip().lower()
        if not email:
            continue
        exists = db.query(Deal).filter(Deal.vendor_email == email,
                                       Deal.mode == e.mode).first()
        if exists:
            continue
        try:
            reply_tracker._record_deal(db, e, [])
            created += 1
        except Exception:
            db.rollback()

    total = db.query(Deal).count()
    return {"created": created, "replied_contacts": len(rows), "total_deals": total,
            "note": "Run a reply check next — it reads each conversation and "
                    "fills in prices, links and turnaround times."}


@router.post("/deals/clean-sites")
def clean_deal_sites(mode: str = "", db: Session = Depends(get_db)):
    """Strip junk out of the site lists already on the deals sheet.

    Removes our own domains, IP addresses, link shorteners and anything that
    isn't really a website — entries that were collected before those checks
    existed.
    """
    from .crm_models import Deal
    from .deal_parser import _valid_domain, _is_ours

    q = db.query(Deal)
    if mode:
        q = q.filter(Deal.mode == mode)

    cleaned = removed = 0
    for d in q.all():
        sites = [x.strip().lower().replace("www.", "")
                 for x in (d.sites or "").split(",") if x.strip()]
        keep, seen = [], set()
        for site in sites:
            if site in seen:
                continue
            # always keep the site we actually contacted
            if site == (d.primary_site or "").lower():
                seen.add(site); keep.append(site); continue
            if not _valid_domain(site) or _is_ours(site):
                removed += 1
                continue
            seen.add(site); keep.append(site)
        primary = (d.primary_site or "").lower()
        if primary and primary in keep:
            keep.remove(primary); keep.insert(0, primary)
        new_val = ",".join(keep)
        if new_val != (d.sites or ""):
            d.sites = new_val
            cleaned += 1
    db.commit()
    return {"rows_cleaned": cleaned, "entries_removed": removed,
            "note": "Our own domains, IPs, shorteners and duplicates are gone."}


@router.get("/deals")
def deals_sheet(mode: str = "", status: str = "", db: Session = Depends(get_db)):
    """The live deals sheet — one row per vendor, filled in from their replies."""
    from .crm_models import Deal
    q = db.query(Deal)
    if mode:
        q = q.filter(Deal.mode == mode)
    if status:
        q = q.filter(Deal.status == status)
    rows = q.order_by(Deal.last_reply_at.desc().nullslast()).all()

    def _fmt(d):
        sites = [x for x in (d.sites or "").split(",") if x]
        return {
            "id": d.id, "mode": d.mode,
            "primary_site": d.primary_site,
            "sites": sites, "site_count": len(sites),
            "vendor_email": d.vendor_email, "our_email": d.our_email,
            "sheet_url": d.sheet_url,
            "currency": d.currency,
            "guest_post_price": d.guest_post_price,
            "link_insert_price": d.link_insert_price,
            "dofollow_links": d.dofollow_links,
            "nofollow_links": d.nofollow_links,
            "tat": d.tat, "sample_url": d.sample_url,
            "status": d.status,
            "deal_date": d.deal_date.strftime("%Y-%m-%d") if d.deal_date else "",
            "last_reply": d.last_reply_at.strftime("%Y-%m-%d %H:%M") if d.last_reply_at else "",
            "notes": d.notes or "",
        }

    out = [_fmt(d) for d in rows]
    return {"deals": out, "total": len(out),
            "dealing": sum(1 for d in out if d["status"] != "done"),
            "done": sum(1 for d in out if d["status"] == "done")}


@router.post("/deals/{deal_id}/update")
def update_deal(deal_id: int, field: str = "", value: str = "",
                db: Session = Depends(get_db)):
    """Edit one cell of the sheet by hand (prices, links, status, notes...)."""
    from .crm_models import Deal
    allowed = {"currency", "guest_post_price", "link_insert_price", "dofollow_links",
               "nofollow_links", "tat", "sample_url", "sheet_url", "status",
               "notes", "primary_site", "our_email"}
    if field not in allowed:
        return {"error": f"'{field}' can't be edited. Allowed: {sorted(allowed)}"}
    d = db.get(Deal, deal_id)
    if not d:
        return {"error": "Deal not found"}
    setattr(d, field, value)
    if field == "status" and value == "done" and not d.deal_date:
        from datetime import datetime as _dt
        d.deal_date = _dt.utcnow()
    db.commit()
    return {"updated": deal_id, "field": field, "value": value}


@router.get("/deals/export")
def deals_export(mode: str = "", format: str = "csv", db: Session = Depends(get_db)):
    """Download the deals sheet as CSV or Excel."""
    from .crm_models import Deal
    import csv, io
    from fastapi.responses import StreamingResponse

    q = db.query(Deal)
    if mode:
        q = q.filter(Deal.mode == mode)
    rows = q.order_by(Deal.last_reply_at.desc().nullslast()).all()

    headers = ["Site", "All sites", "Their email", "Our email", "Price list",
               "Currency", "Guest post", "Link insertion", "Dofollow", "Nofollow",
               "TAT", "Sample", "Status", "Deal date", "Last reply", "Notes"]

    def row_of(d):
        return [d.primary_site or "", d.sites or "", d.vendor_email or "",
                d.our_email or "", d.sheet_url or "", d.currency or "",
                d.guest_post_price or "", d.link_insert_price or "",
                d.dofollow_links or "", d.nofollow_links or "", d.tat or "",
                d.sample_url or "", d.status or "",
                d.deal_date.strftime("%Y-%m-%d") if d.deal_date else "",
                d.last_reply_at.strftime("%Y-%m-%d %H:%M") if d.last_reply_at else "",
                d.notes or ""]

    if format == "xlsx":
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "Deals"
            ws.append(headers)
            for d in rows:
                ws.append(row_of(d))
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return StreamingResponse(
                buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": 'attachment; filename="deals.xlsx"'})
        except Exception:
            pass  # fall through to CSV

    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(headers)
    for d in rows:
        w.writerow(row_of(d))
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": 'attachment; filename="deals.csv"'})


@router.post("/outreach/deal-done")
def mark_deal_done(email: str = "", mode: str = "", done: bool = True,
                   db: Session = Depends(get_db)):
    """Mark a conversation as closed (or reopen it).

    Puts the "Deal Done" label on that thread in Gmail and clears "Dealing".
    """
    from .crm_models import OutreachEntry
    q = db.query(OutreachEntry).filter(
        OutreachEntry.email == (email or "").strip().lower())
    if mode:
        q = q.filter(OutreachEntry.mode == mode)
    rows = q.all()
    if not rows:
        return {"updated": 0, "error": f"No contact found for '{email}'."}

    for e in rows:
        e.deal_stage = "done" if done else ("dealing" if e.status == "replied" else "")
        e.label_target = f"{e.mode}:{'done' if done else 'dealing'}"
        e.label_state = ""            # queue it for re-labelling
    db.commit()
    try:
        from . import gmail_labels
        gmail_labels.kick()
    except Exception:
        pass
    return {"updated": len(rows), "email": email,
            "stage": "done" if done else "reopened",
            "note": "Gmail label updates within a few seconds."}


@router.get("/outreach/deals")
def list_deals(mode: str = "", db: Session = Depends(get_db)):
    """Conversations that are live or closed — the ones worth your attention."""
    from .crm_models import OutreachEntry
    q = db.query(OutreachEntry).filter(
        (OutreachEntry.status == "replied") | (OutreachEntry.deal_stage != ""))
    if mode:
        q = q.filter(OutreachEntry.mode == mode)
    rows = q.order_by(OutreachEntry.replied_at.desc()).limit(500).all()
    return {"deals": [{"email": e.email, "domain": e.domain, "mode": e.mode,
                       "sender": e.sender_email,
                       "stage": (e.deal_stage or ("dealing" if e.status == "replied" else "")),
                       "replied_at": e.replied_at.isoformat() if e.replied_at else ""}
                      for e in rows],
            "dealing": sum(1 for e in rows if (e.deal_stage or "") != "done"),
            "done": sum(1 for e in rows if (e.deal_stage or "") == "done")}


@router.post("/labels/relabel")
def labels_relabel(mode: str = "", db: Session = Depends(get_db)):
    """Re-tag every message we've sent with the label its thread deserves now.

    Works out the right stage for each one — first email, follow-up N, replied
    ("Dealing"), or closed ("Deal Done") — then queues it for the label worker.
    """
    from .crm_models import OutreachEntry

    q = db.query(OutreachEntry).filter(OutreachEntry.message_id != "")
    if mode:
        q = q.filter(OutreachEntry.mode == mode)
    rows = q.all()

    counts = {"base": 0, "followup": 0, "dealing": 0, "done": 0}
    for e in rows:
        if (e.deal_stage or "") == "done":
            stage, bucket = "done", "done"
        elif e.status == "replied":
            stage, bucket = "dealing", "dealing"
        elif (e.followup_count or 0) > 0:
            stage, bucket = str(e.followup_count), "followup"
        else:
            stage, bucket = "0", "base"
        e.label_target = f"{e.mode}:{stage}"
        e.label_state = ""          # queue it
        counts[bucket] += 1
    db.commit()

    # Anything sent before message-ids were recorded can't be found in Gmail
    unlabelable = db.query(OutreachEntry).filter(
        (OutreachEntry.message_id == "") | (OutreachEntry.message_id.is_(None)))
    if mode:
        unlabelable = unlabelable.filter(OutreachEntry.mode == mode)
    missing = unlabelable.count()

    try:
        from . import gmail_labels
        gmail_labels.kick()
    except Exception:
        pass

    return {"queued_for_relabel": len(rows), "breakdown": counts,
            "cannot_label": missing,
            "note": "Labels are applied in the background — a few hundred a minute. "
                    "Messages sent before the app started recording Gmail message "
                    "ids can't be located, so they're skipped."}


@router.get("/labels/status")
def labels_status(db: Session = Depends(get_db)):
    """Per-sender label health: which senders can label, and why not if they can't."""
    from .crm_models import OutreachEntry
    from .database import Sender
    from . import gmail_labels
    waiting = db.query(OutreachEntry).filter(
        OutreachEntry.message_id != "",
        OutreachEntry.label_state != OutreachEntry.label_target).count()

    rows = []
    for s in db.query(Sender).all():
        pend = db.query(OutreachEntry).filter(
            OutreachEntry.sender_email == s.email,
            OutreachEntry.message_id != "",
            OutreachEntry.label_state != OutreachEntry.label_target).count()
        labelled = db.query(OutreachEntry).filter(
            OutreachEntry.sender_email == s.email,
            OutreachEntry.label_state != "",
            OutreachEntry.label_state == OutreachEntry.label_target).count()
        err = gmail_labels._sender_errors.get(s.email, "")
        rows.append({"sender": s.email, "method": s.method,
                     "labelled": labelled, "waiting": pend,
                     "problem": err or ""})

    t = gmail_labels._thread
    return {"worker_running": bool(t is not None and t.is_alive()),
            "waiting_for_label": waiting,
            "senders": rows,
            "labels": list(gmail_labels.LABEL_COLORS.keys())}


@router.post("/followups/test")
def followups_test(email: str = "", mode: str = "vendor", sender: str = "",
                   company: str = "", db: Session = Depends(get_db)):
    """Send ONE sample follow-up to your own address to preview the wording.
    Changes nothing in your real send list."""
    from . import followup_engine
    return followup_engine.send_test(db, to_email=email, mode=mode,
                                     sender_email=sender, company=company)


@router.get("/followups/due")
def followups_due(mode: str = "", delay_hours: int = 24, max_followups: int = 2,
                  db: Session = Depends(get_db)):
    """Preview who is due a reminder — sends nothing."""
    from . import followup_engine
    return followup_engine.run_followups(db, mode=mode, delay_hours=delay_hours,
                                         max_followups=max_followups, dry_run=True)


@router.post("/followups/run")
def followups_run(mode: str = "", delay_hours: int = 30, max_followups: int = 2,
                  limit: int = 60, db: Session = Depends(get_db)):
    """Send reminders now to everyone who hasn't replied."""
    from . import followup_engine
    try:
        return followup_engine.run_followups(db, mode=mode, delay_hours=delay_hours,
                                             max_followups=max_followups, limit=limit)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"sent": 0, "error": str(e)}


@router.post("/followups/auto")
def followups_auto(enabled: bool = True, delay_hours: int = 30, max_followups: int = 2):
    """Turn automatic follow-ups on or off."""
    from . import followup_engine
    if enabled:
        return followup_engine.start(interval_minutes=60, delay_hours=delay_hours,
                                     max_followups=max_followups)
    return followup_engine.stop()


@router.get("/followups/log")
def followups_log(mode: str = "", page: int = 1, limit: int = 100,
                  filter: str = "", db: Session = Depends(get_db)):
    """Every reminder that has gone out — who, from which sender, and what came of it."""
    from .crm_models import OutreachEntry

    q = db.query(OutreachEntry).filter(OutreachEntry.followup_count > 0)
    if mode:
        q = q.filter(OutreachEntry.mode == mode)
    if filter == "replied":
        q = q.filter(OutreachEntry.status == "replied")
    elif filter == "waiting":
        q = q.filter(OutreachEntry.status.in_(["sent", "opened"]))
    elif filter == "1":
        q = q.filter(OutreachEntry.followup_count == 1)
    elif filter == "2":
        q = q.filter(OutreachEntry.followup_count >= 2)

    total = q.count()
    rows = (q.order_by(OutreachEntry.last_followup_at.desc().nullslast())
             .offset((max(1, page) - 1) * limit).limit(limit).all())

    def _fmt(e):
        return {
            "email": e.email, "domain": e.domain,
            "sender": e.sender_email or "",
            "subject": e.subject or "",
            "followup_count": e.followup_count or 0,
            "first_sent": e.sent_at.strftime("%d %b, %H:%M") if e.sent_at else "",
            "last_reminder": e.last_followup_at.strftime("%d %b, %H:%M") if e.last_followup_at else "",
            "status": e.status,
            "replied": e.status == "replied",
        }

    # How many reminders each sender has sent — shows the load is shared
    from sqlalchemy import func
    by_sender_q = db.query(OutreachEntry.sender_email,
                           func.sum(OutreachEntry.followup_count)).filter(
        OutreachEntry.followup_count > 0)
    if mode:
        by_sender_q = by_sender_q.filter(OutreachEntry.mode == mode)
    by_sender = [{"sender": em or "(unknown)", "reminders": int(n or 0)}
                 for em, n in by_sender_q.group_by(OutreachEntry.sender_email).all()]
    by_sender.sort(key=lambda x: -x["reminders"])

    return {"total": total, "page": page, "limit": limit,
            "rows": [_fmt(e) for e in rows], "by_sender": by_sender}


@router.get("/followups/stats")
def followups_stats(mode: str = "", delay_hours: int = 30,
                    db: Session = Depends(get_db)):
    """Live follow-up numbers for the Overview card."""
    from .crm_models import OutreachEntry
    from .database import Sender
    from . import followup_engine
    from datetime import datetime, timedelta

    now = datetime.utcnow()
    q = db.query(OutreachEntry)
    if mode:
        q = q.filter(OutreachEntry.mode == mode)

    awaiting = q.filter(OutreachEntry.status.in_(["sent", "opened"]))
    sent_1 = q.filter(OutreachEntry.followup_count == 1).count()
    sent_2 = q.filter(OutreachEntry.followup_count >= 2).count()

    # How many are past the wait and still owed a reminder
    cutoff = now - timedelta(hours=delay_hours)
    due_now = awaiting.filter(
        OutreachEntry.followup_count < followup_engine.MAX_FOLLOWUPS,
        OutreachEntry.sent_at.isnot(None),
        OutreachEntry.sent_at <= cutoff,
    ).count()

    # Replies that arrived after we nudged — the reason to do this at all
    replied_after = q.filter(OutreachEntry.status == "replied",
                             OutreachEntry.followup_count > 0).count()

    t = followup_engine._thread
    usable_senders = db.query(Sender).filter(
        Sender.status.notin_(["auth_failed", "verifying"]))
    if mode:
        usable_senders = usable_senders.filter(Sender.mode == mode)

    sent_today = followup_engine._sent_today_count(db)

    return {
        "due_now": due_now,
        "reminders_sent": sent_1 + sent_2,
        "first_reminder": sent_1,
        "second_reminder": sent_2,
        "replied_after_reminder": replied_after,
        "waiting_for_reply": awaiting.count(),
        "auto_running": bool(t is not None and t.is_alive()),
        "delay_hours": followup_engine.DEFAULT_DELAY_HOURS,
        "deal_delay_hours": followup_engine.DEAL_DELAY_HOURS,
        "max_per_contact": followup_engine.MAX_FOLLOWUPS,
        "senders": usable_senders.count(),
        # Live rate-limit picture
        "sent_today": sent_today,
        "daily_cap": followup_engine.DAILY_CAP,
        "daily_remaining": max(0, followup_engine.DAILY_CAP - sent_today),
        "per_sender_per_hour": followup_engine.PER_SENDER_PER_HOUR,
    }


@router.get("/followups/status")
def followups_status():
    from . import followup_engine
    t = followup_engine._thread
    return {"running": bool(t is not None and t.is_alive()),
            "delay_hours": followup_engine.DEFAULT_DELAY_HOURS,
            "max_followups": followup_engine.MAX_FOLLOWUPS}


@router.post("/replies/check")
def check_replies_now(days: int = 14, db: Session = Depends(get_db)):
    """Kick off an inbox scan in the background and return immediately.
    Reading several mailboxes takes minutes, so this never blocks the request.
    Poll /crm/replies/status for the result."""
    from . import reply_tracker
    return reply_tracker.check_now()


@router.post("/replies/reset")
def replies_reset(mode: str = "", db: Session = Depends(get_db)):
    """Clear reply marks made by the old address-based matching."""
    from . import reply_tracker
    try:
        return reply_tracker.reset_replies(db, mode=mode)
    except Exception as e:
        import traceback; traceback.print_exc()
        return {"reset": 0, "error": str(e)}


@router.get("/replies/status")
def replies_status():
    """Result of the last reply check + whether one is running right now."""
    from . import reply_tracker
    return reply_tracker.last_result()


@router.post("/blog/check-sites")
def check_blog_sites(sites: str = ""):
    """Pre-check sites before research. Warns about giant portals / aggregators
    (MSN, Yahoo, Forbes...) and unreachable sites so the user doesn't waste time.
    Returns a per-site verdict."""
    from . import blog_research
    from concurrent.futures import ThreadPoolExecutor, wait
    site_list = [s.strip() for s in re.split(r"[,\n]", sites) if s.strip()]
    if not site_list:
        return {"results": [], "total": 0, "warnings": 0, "all_ok": True}

    # Check every site AT THE SAME TIME. Done one by one, ten sites meant up to
    # seventy sequential requests — minutes of staring at a spinner.
    def _one(site):
        try:
            return {"site": site, **blog_research.check_site(site)}
        except Exception as e:
            return {"site": site, "ok": True, "reason": "check_failed",
                    "hint": f"Couldn't pre-check '{site}' ({e}). Research will still try it."}

    # HARD CAP: this is only a sanity check, so it must never hold the user up.
    # Whatever hasn't answered within the budget is simply treated as fine —
    # research will try it anyway.
    # Every site is checked at once, so the whole thing is bounded by the
    # SLOWEST site, not the sum. A generous budget means all sites report in
    # one go — previously slow ones timed out and only surfaced on the next
    # attempt, which felt like errors arriving one at a time.
    BUDGET = 40
    results = []
    ex = ThreadPoolExecutor(max_workers=min(50, len(site_list)))
    try:
        futures = {ex.submit(_one, site): site for site in site_list}
        done, not_done = wait(futures.keys(), timeout=BUDGET)
        for f in done:
            try:
                results.append(f.result())
            except Exception:
                results.append({"site": futures[f], "ok": True, "reason": "check_failed",
                                "hint": ""})
        for f in not_done:
            f.cancel()
            results.append({"site": futures[f], "ok": True, "reason": "check_timeout",
                            "hint": ""})
        if not_done:
            print(f"[BlogResearch] Pre-check: {len(not_done)} site(s) were slow to "
                  f"respond — skipping the check for them.")
    finally:
        ex.shutdown(wait=False)

    # keep the original order so the list reads predictably
    order = {s: i for i, s in enumerate(site_list)}
    results.sort(key=lambda r: order.get(r["site"], 999))

    warnings = sum(1 for r in results if not r.get("ok"))
    return {"results": results, "total": len(site_list),
            "warnings": warnings,
            "all_ok": warnings == 0}


@router.post("/blog/create")
def create_blog_job(name: str = "", sites: str = "", time_range: str = "1m",
                    max_articles: int = 150, autopilot: bool = False,
                    db: Session = Depends(get_db)):
    """Create a blog research job. sites = comma/newline separated domains."""
    from .crm_models import BlogResearchJob
    site_list = [s.strip() for s in re.split(r"[,\n]", sites) if s.strip()]
    if not site_list:
        return {"error": "Add at least one blog site."}
    job = BlogResearchJob(
        name=name or f"Research {datetime.utcnow().strftime('%b %d')}",
        sites=",".join(site_list), time_range=time_range,
        max_articles=max_articles, autopilot=autopilot,
        total_sites=len(site_list), status="pending",
    )
    db.add(job); db.commit(); db.refresh(job)
    return {"job_id": job.id, "name": job.name, "total_sites": len(site_list),
            "time_range": time_range}


@router.post("/blog/{job_id}/start")
def start_blog_job(job_id: int, db: Session = Depends(get_db)):
    """Start researching — find external links across all sites."""
    from .crm_models import BlogResearchJob, BlogResearchLink
    from . import blog_research
    job = db.get(BlogResearchJob, job_id)
    if not job:
        return {"error": "Job not found"}
    # Only refuse if a run is ACTUALLY still going. A leftover entry from a
    # finished or crashed run used to make Start silently do nothing.
    existing = _blog_threads.get(job_id)
    if existing is not None:
        if existing.is_alive():
            return {"error": "Already running"}
        _blog_threads.pop(job_id, None)

    _blog_stop[job_id] = {"stop": False}

    # Flip the status straight away so the UI shows progress immediately instead
    # of sitting on "pending" until the background thread gets going.
    try:
        job.status = "running"
        job.phase = "articles"
        db.commit()
    except Exception:
        db.rollback()

    def _run():
        bg = None
        # Scraper pieces imported up-front so on_link can fire email scraping
        # the instant a link is found (concurrent with the rest of research).
        from . import scraper as _scraper
        from .crm_models import OutreachEntry
        from concurrent.futures import ThreadPoolExecutor as _TPE
        import threading as _threading

        # Collected across every site so we can explain the result afterwards
        run_stats = {}

        # Thread-safe dedup: parallel workers reserve an email here BEFORE the
        # DB insert, so two workers scraping the same email can't both add it.
        _added_emails = set()
        _added_lock = _threading.Lock()

        def _is_junk_email(email, domain):
            """Reject clearly-bad emails like info@wikipedia.can."""
            if not email or "@" not in email:
                return True
            local, _, dom = email.partition("@")
            tld = dom.rsplit(".", 1)[-1] if "." in dom else ""
            bad_tlds = {"can", "con", "cim", "cm", "co m", "comm", "net work",
                        "png", "jpg", "gif", "webp", "svg", "css", "js"}
            if tld.lower() in bad_tlds:
                return True
            if len(tld) < 2 or len(tld) > 10:
                return True
            return False

        def _scrape_one(link_id):
            """Scrape emails for one domain, add real ones to client list.
            Concurrency-safe (own DB session)."""
            s = SessionLocal()
            try:
                lk = s.get(BlogResearchLink, link_id)
                if not lk or lk.email_status != "pending":
                    return
                try:
                    res = _scraper.extract_domain(lk.target_domain)
                    contacts = res.get("contacts", [])
                    email = ""
                    for c in contacts:
                        cand = c.get("email", "")
                        if cand and not _is_junk_email(cand, lk.target_domain):
                            email = cand
                            break
                    if email:
                        email = email.strip().lower()  # normalize for dedupe
                        lk.email = email
                        lk.email_status = "found"
                        from . import compliance as _compliance
                        suppressed = _compliance.is_suppressed(s, email)

                        # Thread-safe reservation: only the first worker to see
                        # this email gets to try inserting it.
                        reserved = False
                        with _added_lock:
                            if email not in _added_emails:
                                _added_emails.add(email)
                                reserved = True

                        if suppressed:
                            print(f"[BlogResearch]   suppressed (skip): {email}")
                        elif not reserved:
                            print(f"[BlogResearch]   dup (skip): {email}")
                        else:
                            # Double-check DB (covers emails added in earlier jobs).
                            # Blog research keeps its OWN list (mode="blog"), separate
                            # from the client dashboard.
                            exists = s.query(OutreachEntry).filter(
                                OutreachEntry.mode == "blog",
                                OutreachEntry.email == email).first()
                            if exists:
                                print(f"[BlogResearch]   dup (skip): {email}")
                            else:
                                s.add(OutreachEntry(
                                    mode="blog", email=email, domain=lk.target_domain,
                                    email_type="blog_research", confidence="medium",
                                    source_url=lk.target_url, status="pending",
                                    ref_site=lk.source_site or "",
                                    ref_article=lk.source_article or ""))
                                print(f"[BlogResearch]   + added {lk.target_domain} -> {email}")
                    else:
                        lk.email_status = "no_email"
                        print(f"[BlogResearch]   no email: {lk.target_domain}")
                    s.commit()
                    jj = s.get(BlogResearchJob, job_id)
                    if jj:
                        jj.emails_found = s.query(BlogResearchLink).filter(
                            BlogResearchLink.job_id == job_id,
                            BlogResearchLink.email_status == "found").count()
                        s.commit()
                except Exception as ex_inner:
                    lk.email_status = "no_email"
                    s.commit()
                    print(f"[BlogResearch]   ! {lk.target_domain} error: {ex_inner}")
            finally:
                s.close()

        # Live email-scrape pool: fires as links come in, parallel to research
        email_pool = _TPE(max_workers=50)   # scrape emails as fast as links are found
        email_futures = []

        try:
            bg = SessionLocal()
            j = bg.get(BlogResearchJob, job_id)
            j.status = "running"; j.done_sites = 0; j.links_found = 0
            j.articles_found = 0; j.emails_found = 0; j.phase = "articles"
            bg.commit()
            sites = [s for s in j.sites.split(",") if s]
            # Tell the extractor which domains we're researching, so a link
            # pointing back at one of them is never offered as a prospect.
            try:
                blog_research.set_source_sites(sites)
            except Exception:
                pass

            # And our OWN domains — the company site plus every sender's domain.
            # These appear in signatures and footers constantly; they are not leads.
            try:
                from .database import Sender as _S
                from .config import settings as _cfg
                own = set()
                for snd in bg.query(_S).all():
                    if snd.email and "@" in snd.email:
                        own.add(snd.email.split("@")[-1].lower())
                for extra in (getattr(_cfg, "SENDING_DOMAIN", ""),
                              getattr(_cfg, "PUBLIC_URL", "")):
                    if extra:
                        own.add(extra)
                blog_research.set_own_domains(own)
                if own:
                    print(f"[BlogResearch] Ignoring our own domains: "
                          f"{', '.join(sorted(d for d in blog_research._OWN_DOMAINS))}")
            except Exception as oe:
                print(f"[BlogResearch] Could not read our own domains: {oe}")
            seen_domains = set()  # global dedupe across all sites

            for site in sites:
                if _blog_stop.get(job_id, {}).get("stop"):
                    break

                # Live callback: article opened
                def on_article(article_url, count):
                    jj = bg.get(BlogResearchJob, job_id)
                    if jj:
                        jj.articles_found = (jj.articles_found or 0) + 1
                        jj.phase = "articles"
                        bg.commit()

                # Live callback: new link found -> save + fire email scrape NOW
                def on_link(r):
                    if r["target_domain"] in seen_domains:
                        return
                    seen_domains.add(r["target_domain"])
                    link = BlogResearchLink(
                        job_id=job_id, source_site=r["source_site"],
                        source_article=r["source_article"],
                        target_domain=r["target_domain"], target_url=r["target_url"],
                        published_date=r.get("published_date", ""),
                        category=r.get("category", ""),
                    )
                    bg.add(link)
                    bg.flush()  # assign link.id without a full commit
                    new_link_id = link.id
                    jj = bg.get(BlogResearchJob, job_id)
                    if jj:
                        jj.links_found = (jj.links_found or 0) + 1
                        jj.phase = "links"
                    bg.commit()
                    # AUTO email scrape: fire this domain into the pool immediately
                    # so email scraping runs concurrently with the research crawl.
                    if not _blog_stop.get(job_id, {}).get("stop"):
                        email_futures.append(email_pool.submit(_scrape_one, new_link_id))

                try:
                    print(f"[BlogResearch] Researching {site}...")
                    blog_research.research_site(
                        site, j.time_range, j.max_articles,
                        workers=50, on_article=on_article, on_link=on_link,
                        should_stop=lambda: _blog_stop.get(job_id, {}).get("stop", False),
                        on_stats=run_stats)
                    print(f"[BlogResearch] Done {site}")
                except Exception as e:
                    import traceback
                    print(f"[BlogResearch] {site} ERROR: {e}")
                    traceback.print_exc()

                jj = bg.get(BlogResearchJob, job_id)
                jj.done_sites = (jj.done_sites or 0) + 1
                bg.commit()

            j = bg.get(BlogResearchJob, job_id)
            j.status = "done"; j.phase = "emails"
            j.links_found = bg.query(BlogResearchLink).filter(
                BlogResearchLink.job_id == job_id).count()
            # Explain the outcome in plain words, so "0 links" is never a mystery
            try:
                bits = []
                if run_stats.get("ok"):          bits.append(f"{run_stats['ok']} article(s) had outbound links")
                if run_stats.get("no_links"):    bits.append(f"{run_stats['no_links']} had none")
                if run_stats.get("too_old"):     bits.append(f"{run_stats['too_old']} were outside the {j.time_range} window")
                if run_stats.get("no_date"):     bits.append(f"{run_stats['no_date']} had no confirmable date (skipped)")
                if run_stats.get("unreachable"): bits.append(f"{run_stats['unreachable']} wouldn't load")
                if run_stats.get("blocked"):     bits.append(f"{run_stats['blocked']} were blocked by the site")
                # If nothing was found, say what was actually on those pages —
                # otherwise "0 links" tells the user nothing they can act on.
                ex = run_stats.get("no_link_examples") or []
                if ex and not run_stats.get("ok"):
                    e0 = ex[0]
                    if e0.get("external", 0) > 0:
                        bits.append(f"pages had {e0['external']} outbound link(s) but "
                                    f"all were filtered out")
                    elif e0.get("anchors", 0) == 0:
                        bits.append("article pages came back empty (likely rendered "
                                    "by JavaScript, or bot-blocked)")
                    else:
                        bits.append(f"articles had {e0.get('anchors', 0)} link(s), none "
                                    f"pointing to other sites")
                j.summary = " · ".join(bits)
                if j.summary:
                    print(f"[BlogResearch] Job {job_id} summary: {j.summary}")
            except Exception:
                pass
            bg.commit()
            print(f"[BlogResearch] Job {job_id} research complete: {j.links_found} links. "
                  f"Waiting for {len(email_futures)} live email scrapes...")

            # Wait for the live email scrapes, but with an OVERALL deadline
            # rather than up to 60s EACH in submission order. Waiting per-future
            # in order meant a few slow domains could hold the whole job for many
            # minutes (or effectively forever) even though most were long done —
            # that's what left the job stuck at "Extracting links..." with the
            # site counter frozen. as_completed drains them as they finish and we
            # stop waiting past a sane cap; anything unfinished is left pending
            # and mopped up (bounded) below.
            from concurrent.futures import as_completed as _ascomp
            deadline = time.time() + 120       # at most 2 min for the live batch
            try:
                for f in _ascomp(email_futures, timeout=120):
                    try:
                        f.result(timeout=1)
                    except Exception:
                        pass
                    if time.time() > deadline:
                        break
            except Exception:
                pass  # as_completed itself timed out — fall through to cleanup

            # Safety net: any domain that never got scraped (e.g. fired after a
            # stop check) gets picked up here.
            leftover = bg.query(BlogResearchLink).filter(
                BlogResearchLink.job_id == job_id,
                BlogResearchLink.email_status == "pending").all()
            leftover_ids = [l.id for l in leftover]
            if leftover_ids:
                print(f"[BlogResearch] Scraping {len(leftover_ids)} leftover domains...")
                # Fire them all into the pool at once and drain with an overall
                # deadline. The old code did .result(timeout=60) one domain at a
                # time — 40 slow leftovers meant 40 minutes serial. Now it's
                # parallel and bounded.
                from concurrent.futures import as_completed as _ascomp2
                lf = [email_pool.submit(_scrape_one, lid) for lid in leftover_ids]
                try:
                    for f in _ascomp2(lf, timeout=180):
                        try:
                            f.result(timeout=1)
                        except Exception:
                            pass
                except Exception:
                    pass

            # Final count + mark fully done
            jj = bg.get(BlogResearchJob, job_id)
            if jj:
                jj.emails_found = bg.query(BlogResearchLink).filter(
                    BlogResearchLink.job_id == job_id,
                    BlogResearchLink.email_status == "found").count()
                jj.phase = "done"
                bg.commit()
            print(f"[BlogResearch] Job {job_id} FULLY done: {jj.emails_found if jj else 0} emails found")

            # ============ AUTOPILOT: auto-send if enabled ============
            # If this job was created with autopilot=ON, automatically create a
            # client campaign from the freshly-scraped emails and start sending —
            # no manual button needed. Research -> scrape -> send, fully hands-off.
            if jj and jj.autopilot:
                try:
                    from .crm_models import OutreachEntry
                    from .database import Sender
                    from . import send_engine

                    # Count pending BLOG emails ready to send (blog has its own list)
                    pending_count = bg.query(OutreachEntry).filter(
                        OutreachEntry.mode == "blog",
                        OutreachEntry.status == "pending").count()

                    # Senders: blog reuses the CLIENT senders (blog pitches clients).
                    # Fall back to any active sender if none are tagged client.
                    active_senders = bg.query(Sender).filter(
                        Sender.mode == "client",
                        Sender.status.notin_(["auth_failed", "verifying"])
                    ).all()
                    if not active_senders:
                        active_senders = bg.query(Sender).filter(
                            Sender.status.notin_(["auth_failed", "verifying"])
                        ).all()

                    if pending_count > 0 and active_senders:
                        camp = Campaign(
                            name=f"Autopilot — {jj.name} ({pending_count} emails)",
                            mode="blog",
                            status="ready",
                            sender_emails=",".join(s.email for s in active_senders),
                            emails_per_batch=10,
                            delay_seconds=30,
                            min_delay_sec=15,
                            max_delay_sec=40,
                            autopilot=True,
                            total_target=pending_count,
                        )
                        bg.add(camp)
                        bg.commit()
                        bg.refresh(camp)
                        print(f"[BlogResearch] Autopilot: created campaign #{camp.id}, "
                              f"starting send of {pending_count} emails via "
                              f"{len(active_senders)} sender(s)...")
                        # Kick off the send engine (runs in its own thread)
                        send_engine.start_campaign_send(
                            campaign_id=camp.id, mode="blog",
                            emails_per_batch=10, delay_seconds=30,
                            min_delay=15, max_delay=40,
                            sender_filter=camp.sender_emails,
                            total_target=pending_count,
                            autopilot=True,
                        )
                    else:
                        reason = ("no pending emails" if pending_count == 0
                                  else "no active senders")
                        print(f"[BlogResearch] Autopilot skipped: {reason}. "
                              f"Emails saved to blog list — send manually.")
                except Exception as ap_err:
                    print(f"[BlogResearch] Autopilot error: {ap_err}")
        except Exception as e:
            import traceback
            print(f"[BlogResearch] Job {job_id} FATAL ERROR: {e}")
            traceback.print_exc()
            # Mark job as error so it doesn't stay stuck on "pending"
            try:
                if bg is None:
                    bg = SessionLocal()
                jj = bg.get(BlogResearchJob, job_id)
                if jj:
                    jj.status = "error"
                    bg.commit()
            except Exception:
                pass
        finally:
            # Shut down the live email-scrape pool (don't leak threads)
            try:
                email_pool.shutdown(wait=False, cancel_futures=True)
            except Exception:
                try:
                    email_pool.shutdown(wait=False)
                except Exception:
                    pass
            if bg is not None:
                bg.close()
            _blog_threads.pop(job_id, None)
            _blog_stop.pop(job_id, None)

    import threading
    t = threading.Thread(target=_run, daemon=True)
    _blog_threads[job_id] = t
    t.start()
    return {"status": "started", "job_id": job_id}


@router.post("/blog/{job_id}/stop")
def stop_blog_job(job_id: int, db: Session = Depends(get_db)):
    """Immediately stop a running blog research job."""
    from .crm_models import BlogResearchJob
    # 1) Set the stop flag so the background thread halts at its next check
    if job_id in _blog_stop:
        _blog_stop[job_id]["stop"] = True
    # 2) Immediately reflect the stop in the DB so the UI updates on next poll
    job = db.get(BlogResearchJob, job_id)
    if job and job.status == "running":
        job.status = "stopped"
        job.phase = ""
        db.commit()
    return {"status": "stopped"}


@router.get("/blog/jobs")
def list_blog_jobs(db: Session = Depends(get_db)):
    from .crm_models import BlogResearchJob
    jobs = db.query(BlogResearchJob).order_by(BlogResearchJob.id.desc()).all()
    return [{"id": j.id, "name": j.name, "status": j.status,
             "total_sites": j.total_sites, "done_sites": j.done_sites,
             "articles_found": j.articles_found or 0,
             "links_found": j.links_found, "emails_found": j.emails_found or 0,
             "phase": j.phase or "", "time_range": j.time_range,
             "summary": getattr(j, "summary", "") or "",
             "autopilot": j.autopilot,
             "created_at": j.created_at.isoformat() if j.created_at else ""} for j in jobs]


@router.get("/blog/{job_id}/links")
def blog_links(job_id: int, page: int = 1, limit: int = 100, db: Session = Depends(get_db)):
    from .crm_models import BlogResearchLink
    q = db.query(BlogResearchLink).filter(BlogResearchLink.job_id == job_id)
    total = q.count()
    links = q.order_by(BlogResearchLink.id.desc()).offset((page-1)*limit).limit(limit).all()
    return {"total": total, "page": page,
            "links": [{"id": l.id, "source_site": l.source_site,
                       "source_article": l.source_article,
                       "target_domain": l.target_domain, "target_url": l.target_url,
                       "email": l.email, "email_status": l.email_status,
                       "published_date": getattr(l, "published_date", ""),
                       "category": getattr(l, "category", "")} for l in links]}


@router.get("/blog/{job_id}/export")
def blog_export(job_id: int, format: str = "csv", only_emails: bool = False,
                db: Session = Depends(get_db)):
    """Export a blog research job's discovered links/prospects as CSV or Excel.
    only_emails=True -> just the rows where an email was found."""
    from .crm_models import BlogResearchLink, BlogResearchJob
    from fastapi.responses import Response

    job = db.get(BlogResearchJob, job_id)
    jobname = (job.name if job else f"job{job_id}").replace(" ", "_").replace("/", "-")

    q = db.query(BlogResearchLink).filter(BlogResearchLink.job_id == job_id)
    if only_emails:
        q = q.filter(BlogResearchLink.email_status == "found")
    links = q.order_by(BlogResearchLink.id).all()

    headers_row = ["From Site", "Category", "Published Date", "Source Article",
                   "Target Domain", "Target URL", "Email", "Email Status"]

    def row_of(l):
        return [l.source_site or "", getattr(l, "category", "") or "",
                getattr(l, "published_date", "") or "",
                l.source_article or "", l.target_domain or "",
                l.target_url or "", l.email or "", l.email_status or ""]

    if format == "xlsx":
        from openpyxl import Workbook
        from io import BytesIO
        wb = Workbook(); ws = wb.active; ws.title = "Prospects"
        ws.append(headers_row)
        for l in links:
            ws.append(row_of(l))
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=blog_{jobname}.xlsx"})

    # CSV
    import csv, io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers_row)
    for l in links:
        w.writerow(row_of(l))
    return Response(content=buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=blog_{jobname}.csv"})



    """Scrape emails from all discovered link domains (like the main scraper)."""
    from .crm_models import BlogResearchLink, OutreachEntry
    from . import scraper
    links = db.query(BlogResearchLink).filter(
        BlogResearchLink.job_id == job_id,
        BlogResearchLink.email_status == "pending").all()

    def _run():
        bg = SessionLocal()
        try:
            for link in bg.query(BlogResearchLink).filter(
                    BlogResearchLink.job_id == job_id,
                    BlogResearchLink.email_status == "pending").all():
                try:
                    result = scraper.extract_domain(link.target_domain, mode="client")
                    contacts = result.get("contacts", [])
                    if contacts:
                        email = (contacts[0].get("email", "") or "").strip().lower()
                        link.email = email
                        link.email_status = "found"
                        # Add to the BLOG send list (not client — blog keeps its own).
                        # Duplicate check must be mode-scoped too.
                        existing = bg.query(OutreachEntry).filter(
                            OutreachEntry.mode == "blog",
                            OutreachEntry.email == email).first()
                        if not existing and email:
                            bg.add(OutreachEntry(
                                mode="blog", email=email, domain=link.target_domain,
                                email_type="blog_research", confidence="medium",
                                source_url=link.target_url, status="pending",
                                ref_site=link.source_site or "",
                                ref_article=link.source_article or ""))
                    else:
                        link.email_status = "no_email"
                    bg.commit()
                except Exception as e:
                    link.email_status = "no_email"
                    bg.commit()
        finally:
            bg.close()

    import threading
    threading.Thread(target=_run, daemon=True).start()
    return {"status": "scraping", "to_scrape": len(links)}


@router.delete("/blog/{job_id}")
def delete_blog_job(job_id: int, db: Session = Depends(get_db)):
    from .crm_models import BlogResearchJob, BlogResearchLink
    job = db.get(BlogResearchJob, job_id)
    if job:
        db.query(BlogResearchLink).filter(BlogResearchLink.job_id == job_id).delete()
        db.delete(job); db.commit()
    return {"deleted": job_id}


@router.post("/blog/migrate-from-client")
def blog_migrate_from_client(db: Session = Depends(get_db)):
    """One-time migration: move OLD blog-research emails that were saved under
    mode='client' (before blog got its own list) into mode='blog'. Skips any
    email that already exists in the blog list (no duplicates)."""
    from .crm_models import OutreachEntry

    # All client-mode entries that came from blog research
    old = db.query(OutreachEntry).filter(
        OutreachEntry.mode == "client",
        OutreachEntry.email_type == "blog_research").all()

    # Emails already present in the blog list (to avoid duplicates)
    existing_blog = {e.email for e in db.query(OutreachEntry).filter(
        OutreachEntry.mode == "blog").all()}

    moved = skipped = 0
    for e in old:
        if e.email in existing_blog:
            # Already in blog list — remove the duplicate client copy
            db.delete(e)
            skipped += 1
        else:
            e.mode = "blog"
            existing_blog.add(e.email)
            moved += 1
    db.commit()
    return {"moved": moved, "skipped_duplicates": skipped,
            "message": f"Moved {moved} blog emails to blog dashboard, "
                       f"removed {skipped} duplicates."}


@router.get("/blog/debug")
def blog_debug(site: str = "techbullion.com", time_range: str = "1m"):
    """Diagnostic: test what the engine sees for a given site."""
    from . import blog_research
    from datetime import datetime
    out = {"site": site, "time_range": time_range}
    root = blog_research._root(site)
    out["root"] = root
    base = "https://" + root
    html = blog_research._fetch(base)
    if not html:
        html = blog_research._fetch("http://" + root)
    out["homepage_fetched"] = bool(html)
    out["homepage_size"] = len(html) if html else 0

    # --- Sitemap diagnostics: which sitemap works, how many entries, date impact ---
    delta = blog_research.TIME_RANGES.get(time_range, blog_research.TIME_RANGES["1m"])
    cutoff = datetime.utcnow() - delta
    out["cutoff_date"] = cutoff.strftime("%Y-%m-%d")
    sitemap_urls = ["/post-sitemap.xml", "/post-sitemap1.xml", "/sitemap-posts.xml",
                    "/wp-sitemap-posts-post-1.xml", "/sitemap_index.xml",
                    "/sitemap.xml", "/wp-sitemap.xml", "/news-sitemap.xml"]
    sitemap_report = []
    for sm in sitemap_urls:
        sm_html = blog_research._fetch(base + sm)
        if not sm_html:
            sitemap_report.append({"path": sm, "found": False})
            continue
        entries = blog_research._parse_sitemap_entries(sm_html)
        subs = [u for u, _ in entries if u.endswith(".xml")]
        # How many would survive the date filter?
        dated = with_date = too_old = 0
        for u, mod in entries[:200]:
            d = blog_research._parse_date(mod) or blog_research._date_from_url(u)
            if d:
                with_date += 1
                if d < cutoff:
                    too_old += 1
            dated += 1
        rep = {
            "path": sm, "found": True, "size": len(sm_html),
            "total_entries": len(entries),
            "sub_sitemaps": len(subs), "sample_sub": subs[:2],
            "checked": dated, "had_date": with_date, "dropped_too_old": too_old,
            "sample_dates": [m for _, m in entries[:3]],
        }
        # If parsing found nothing, show what the content actually looks like
        if len(entries) == 0:
            rep["raw_preview"] = sm_html[:400]
            rep["has_loc_tag"] = "<loc" in sm_html.lower()
            rep["has_url_tag"] = "<url" in sm_html.lower()
            rep["looks_like_html"] = ("<!doctype html" in sm_html[:200].lower()
                                      or "<html" in sm_html[:200].lower())
        sitemap_report.append(rep)
        # stop after first working non-index sitemap with real entries
        if entries and not subs:
            break
    out["sitemaps"] = sitemap_report

    try:
        articles = blog_research._find_articles(site, 150, time_range)
        out["articles_found"] = len(articles)
        out["sample_articles"] = articles[:5]
    except Exception as e:
        import traceback
        out["articles_error"] = str(e)
        out["articles_traceback"] = traceback.format_exc()[-500:]
        articles = []
    # Check links in first 3 articles (not just 1)
    total_links = 0
    article_links = []
    for art in articles[:3]:
        art_url = art[0] if isinstance(art, (list, tuple)) else art
        art_date = art[1] if isinstance(art, (list, tuple)) and len(art) > 1 else ""
        try:
            links, reason = blog_research._extract_external_links(art_url)
            total_links += len(links)
            # DEEP DIAGNOSTIC: count raw anchors vs external ones to see where
            # links are being lost (no body container? all internal? filtered?)
            diag = {}
            try:
                from bs4 import BeautifulSoup
                html = blog_research._fetch(art_url)
                if html:
                    diag["html_size"] = len(html)
                    soup = BeautifulSoup(html, "html.parser")
                    all_a = soup.find_all("a", href=True)
                    diag["total_anchors"] = len(all_a)
                    ext = [a for a in all_a if a["href"].startswith("http")
                           and blog_research._root(a["href"]) != blog_research._root(art_url)]
                    diag["external_anchors"] = len(ext)
                    diag["sample_external"] = [a["href"][:60] for a in ext[:5]]
                    # Which body container was detected?
                    import re as _re
                    body = (soup.find("article") or
                            soup.find("div", class_=_re.compile(r"(entry-content|post-content|article-content|article-body|story-body|content)", _re.I)) or
                            soup.find("main"))
                    diag["body_container"] = (body.name + "." + " ".join(body.get("class", []))) if body else "NONE (using whole page)"
            except Exception as de:
                diag["diag_error"] = str(de)
            article_links.append({"article": art_url, "date": art_date,
                                   "links": len(links), "reason": reason,
                                   "domains": [l[0] for l in links[:8]],
                                   "diag": diag})
        except Exception as e:
            article_links.append({"article": art_url, "error": str(e)})
    out["total_links_in_first_3"] = total_links
    out["per_article"] = article_links
    return out
